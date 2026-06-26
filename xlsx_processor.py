import os
import logging
import openpyxl
from typing import List
from ner_engine import extract_all_entities
from docx_processor import chunk_text

logger = logging.getLogger("docguard")

def process_xlsx(input_path: str, output_path: str):
    logger.info(f"  ▶ 正在处理 Excel 文档: {os.path.basename(input_path)}")
    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        logger.error(f"  ❌ 读取 {input_path} 失败: {e}")
        return False, ""
        
    # 1. 提取所有单元格文本
    full_text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.strip():
                        full_text.append(cell.value)
                        
    raw_text = "\n".join(full_text)
    if not raw_text.strip():
        logger.warning(f"  [警告] {os.path.basename(input_path)} 未提取到文本。")
        return False, ""
        
    # 2. 分块送给大模型进行实体识别
    text_chunks = chunk_text(raw_text)
    entities = extract_all_entities(text_chunks)
    logger.info(f"    ✅ 共找到 {len(entities)} 个敏感实体: {entities[:5]}...")
    
    # 3. 遍历单元格，执行精确替换
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_val = cell.value
                    for entity in entities:
                        if entity in new_val:
                            new_val = new_val.replace(entity, "[脱敏]")
                    cell.value = new_val
                    
    # 4. 保存
    wb.save(output_path)
    
    clean_text = raw_text
    for entity in entities:
        clean_text = clean_text.replace(entity, "[脱敏]")
        
    return True, clean_text

from ner_engine import extract_entities_with_types
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from reversible.placeholder_registry import GlobalMapping


def process_xlsx_reversible(input_path: str, output_path: str, mapping: "GlobalMapping") -> tuple[bool, str]:
    logger.info(f"  ▶ 正在可逆脱敏 Excel 文档: {os.path.basename(input_path)}")
    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        logger.error(f"  ❌ 读取 {input_path} 失败: {e}")
        return False, ""

    # 提取所有单元格文本拼接
    full_text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    full_text.append(cell.value)

    raw_text = "\n".join(full_text)
    if not raw_text.strip():
        logger.warning(f"  [警告] {os.path.basename(input_path)} 未提取到文本。")
        return False, ""

    # 分块调用带类型的实体识别
    text_chunks = chunk_text(raw_text)
    entities_with_types = extract_entities_with_types(text_chunks)
    logger.info(f"    ✅ 共找到 {len(entities_with_types)} 个实体（含类型）")

    # 对每个实体分配占位符
    sorted_entities = [
        (original, mapping.get_or_create(original, entity_type), entity_type)
        for original, entity_type in entities_with_types
    ]
    # entities_with_types 已按长度降序排列，sorted_entities 顺序一致

    # 遍历单元格，对 str 类型单元格执行字符串替换
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_val = cell.value
                    for original, placeholder, _ in sorted_entities:
                        if original in new_val:
                            new_val = new_val.replace(original, placeholder)
                    cell.value = new_val

    wb.save(output_path)

    # 生成脱敏后的 clean text
    anonymized_clean_text = raw_text
    for original, placeholder, _ in sorted_entities:
        anonymized_clean_text = anonymized_clean_text.replace(original, placeholder)

    return True, anonymized_clean_text
